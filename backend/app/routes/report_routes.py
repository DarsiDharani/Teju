"""
Report Routes Module

Purpose: API routes for generating comprehensive reports for admin dashboard
Features:
- User activity reports
- Training performance reports
- Feedback analysis reports
- Skills gap analysis reports
- System usage reports
- Excel export with multiple sheets
- Chart data aggregation for various visualizations

@author Orbit Skill Development Team
@date 2025
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func, or_, and_, desc, case, distinct
from sqlalchemy.orm import selectinload
from typing import Optional, List, Dict, Any
from datetime import datetime, date, timedelta
from io import BytesIO
import json

from app.database import get_db_async
from app.auth_utils import get_current_active_admin
from app.models import (
    User, ManagerEmployee, EmployeeCompetency, AdditionalSkill,
    TrainingDetail, TrainingAssignment, TrainingRequest, TrainingAttendance,
    Trainer, AssignmentSubmission, FeedbackSubmission, Notification
)
from pydantic import BaseModel

router = APIRouter(prefix="/reports", tags=["Reports"])

# ==================== SCHEMAS ====================

class ChartDataResponse(BaseModel):
    """Response model for chart data"""
    labels: List[str]
    datasets: List[Dict[str, Any]]
    chartType: str
    title: str

class ReportSummaryResponse(BaseModel):
    """Response model for report summary statistics"""
    total_count: int
    metrics: Dict[str, Any]
    chartData: Optional[ChartDataResponse] = None

# ==================== HELPER FUNCTIONS ====================

async def get_employee_name(db: AsyncSession, empid: str) -> str:
    """Get employee name from manager_employee table"""
    result = await db.execute(
        select(ManagerEmployee).where(
            or_(
                ManagerEmployee.employee_empid == empid,
                ManagerEmployee.manager_empid == empid
            )
        ).limit(1)
    )
    emp_row = result.scalars().first()
    if emp_row:
        return emp_row.employee_name if emp_row.employee_empid == empid else emp_row.manager_name
    return empid

# ==================== ENDPOINTS ====================

@router.get("/users-overview")
async def get_users_overview_report(
    db: AsyncSession = Depends(get_db_async),
    current_admin = Depends(get_current_active_admin)
):
    """
    Get comprehensive user overview report with statistics
    - Total users by role
    - Active vs inactive users
    - User registration trends
    - Trainer distribution
    """
    try:
        # Get all users with manager relationships
        users_result = await db.execute(select(User))
        all_users = users_result.scalars().all()
        
        # Get manager-employee relationships
        me_result = await db.execute(select(ManagerEmployee))
        me_relationships = me_result.scalars().all()
        
        # Build role mapping
        role_map = {}
        trainer_map = {}
        for rel in me_relationships:
            # Managers
            if rel.manager_empid not in role_map:
                role_map[rel.manager_empid] = 'manager'
            trainer_map[rel.manager_empid] = rel.manager_is_trainer
            
            # Employees
            if rel.employee_empid not in role_map:
                role_map[rel.employee_empid] = 'employee'
            trainer_map[rel.employee_empid] = rel.employee_is_trainer
        
        # Count statistics
        total_users = len(all_users)
        managers_count = sum(1 for role in role_map.values() if role == 'manager')
        employees_count = total_users - managers_count
        trainers_count = sum(1 for is_trainer in trainer_map.values() if is_trainer)
        
        # Registration trend (last 6 months)
        six_months_ago = datetime.utcnow() - timedelta(days=180)
        registrations_by_month = {}
        for user in all_users:
            if user.created_at and user.created_at >= six_months_ago:
                month_key = user.created_at.strftime('%Y-%m')
                registrations_by_month[month_key] = registrations_by_month.get(month_key, 0) + 1
        
        # Sort months
        sorted_months = sorted(registrations_by_month.keys())
        
        return {
            "total_count": total_users,
            "metrics": {
                "total_users": total_users,
                "managers": managers_count,
                "employees": employees_count,
                "trainers": trainers_count,
                "trainer_percentage": round((trainers_count / total_users * 100), 2) if total_users > 0 else 0
            },
            "chartData": {
                "labels": ["Managers", "Employees", "Trainers"],
                "datasets": [{
                    "label": "User Distribution",
                    "data": [managers_count, employees_count, trainers_count],
                    "backgroundColor": ["#3B82F6", "#10B981", "#F59E0B"]
                }],
                "chartType": "doughnut",
                "title": "User Distribution by Role"
            },
            "registrationTrend": {
                "labels": sorted_months,
                "data": [registrations_by_month[m] for m in sorted_months],
                "chartType": "line",
                "title": "User Registration Trend"
            }
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating users overview report: {str(e)}"
        )

@router.get("/trainings-performance")
async def get_trainings_performance_report(
    db: AsyncSession = Depends(get_db_async),
    current_admin = Depends(get_current_active_admin)
):
    """
    Get training performance metrics
    - Training completion rates
    - Attendance statistics
    - Training by department/division
    - Most popular trainings
    """
    try:
        # Get all trainings with assignments and attendance
        trainings_result = await db.execute(
            select(TrainingDetail).options(
                selectinload(TrainingDetail.assignments),
                selectinload(TrainingDetail.attendance_records)
            )
        )
        trainings = trainings_result.scalars().all()
        
        # Calculate statistics
        total_trainings = len(trainings)
        total_assignments = 0
        total_attended = 0
        trainings_by_dept = {}
        trainings_by_skill = {}
        top_trainings = []
        
        for training in trainings:
            assigned_count = len(training.assignments)
            attended_count = sum(1 for att in training.attendance_records if att.attended)
            
            total_assignments += assigned_count
            total_attended += attended_count
            
            # By department
            dept = training.department or "Unknown"
            if dept not in trainings_by_dept:
                trainings_by_dept[dept] = {"count": 0, "assigned": 0, "attended": 0}
            trainings_by_dept[dept]["count"] += 1
            trainings_by_dept[dept]["assigned"] += assigned_count
            trainings_by_dept[dept]["attended"] += attended_count
            
            # By skill
            skill = training.skill or "General"
            trainings_by_skill[skill] = trainings_by_skill.get(skill, 0) + 1
            
            # Top trainings
            completion_rate = (attended_count / assigned_count * 100) if assigned_count > 0 else 0
            top_trainings.append({
                "name": training.training_name,
                "assigned": assigned_count,
                "attended": attended_count,
                "completion_rate": round(completion_rate, 2)
            })
        
        # Sort and get top 10
        top_trainings.sort(key=lambda x: x["assigned"], reverse=True)
        top_trainings = top_trainings[:10]
        
        overall_completion_rate = (total_attended / total_assignments * 100) if total_assignments > 0 else 0
        
        # Department chart data
        dept_labels = list(trainings_by_dept.keys())
        dept_counts = [trainings_by_dept[d]["count"] for d in dept_labels]
        
        # Skill distribution chart data
        skill_labels = list(trainings_by_skill.keys())[:10]  # Top 10 skills
        skill_counts = [trainings_by_skill[s] for s in skill_labels]
        
        return {
            "total_count": total_trainings,
            "metrics": {
                "total_trainings": total_trainings,
                "total_assignments": total_assignments,
                "total_attended": total_attended,
                "overall_completion_rate": round(overall_completion_rate, 2),
                "departments_count": len(trainings_by_dept),
                "skills_count": len(trainings_by_skill)
            },
            "departmentChart": {
                "labels": dept_labels,
                "datasets": [{
                    "label": "Trainings by Department",
                    "data": dept_counts,
                    "backgroundColor": ["#3B82F6", "#10B981", "#F59E0B", "#EF4444", "#8B5CF6", "#EC4899"]
                }],
                "chartType": "bar",
                "title": "Trainings by Department"
            },
            "skillChart": {
                "labels": skill_labels,
                "datasets": [{
                    "label": "Trainings by Skill",
                    "data": skill_counts,
                    "backgroundColor": "#3B82F6"
                }],
                "chartType": "bar",
                "title": "Top Skills Trained"
            },
            "topTrainings": top_trainings
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating trainings performance report: {str(e)}"
        )

@router.get("/feedback-analysis")
async def get_feedback_analysis_report(
    db: AsyncSession = Depends(get_db_async),
    current_admin = Depends(get_current_active_admin)
):
    """
    Get feedback submission and rating analysis
    - Feedback submission rates
    - Average ratings by training
    - Feedback trends over time
    - Response rate analysis
    """
    try:
        # Get all feedback submissions
        feedback_result = await db.execute(
            select(FeedbackSubmission).options(
                selectinload(FeedbackSubmission.training)
            )
        )
        feedbacks = feedback_result.scalars().all()
        
        # Get total assignments for response rate calculation
        assignments_result = await db.execute(
            select(TrainingAssignment)
        )
        assignments = assignments_result.scalars().all()
        
        total_feedbacks = len(feedbacks)
        total_assignments = len(assignments)
        response_rate = (total_feedbacks / total_assignments * 100) if total_assignments > 0 else 0
        
        # Feedback by training
        feedback_by_training = {}
        feedback_by_month = {}
        
        for feedback in feedbacks:
            # By training
            training_id = feedback.training_id
            if training_id not in feedback_by_training:
                feedback_by_training[training_id] = {
                    "training_name": feedback.training.training_name if feedback.training else "Unknown",
                    "count": 0,
                    "submitted_at_list": []
                }
            feedback_by_training[training_id]["count"] += 1
            if feedback.submitted_at:
                feedback_by_training[training_id]["submitted_at_list"].append(feedback.submitted_at)
            
            # By month
            if feedback.submitted_at:
                month_key = feedback.submitted_at.strftime('%Y-%m')
                feedback_by_month[month_key] = feedback_by_month.get(month_key, 0) + 1
        
        # Top trainings by feedback count
        top_feedback_trainings = []
        for tid, data in feedback_by_training.items():
            top_feedback_trainings.append({
                "training_name": data["training_name"],
                "feedback_count": data["count"]
            })
        top_feedback_trainings.sort(key=lambda x: x["feedback_count"], reverse=True)
        top_feedback_trainings = top_feedback_trainings[:10]
        
        # Monthly trend
        sorted_months = sorted(feedback_by_month.keys())
        
        return {
            "total_count": total_feedbacks,
            "metrics": {
                "total_feedbacks": total_feedbacks,
                "total_assignments": total_assignments,
                "response_rate": round(response_rate, 2),
                "unique_trainings_with_feedback": len(feedback_by_training)
            },
            "feedbackTrendChart": {
                "labels": sorted_months,
                "datasets": [{
                    "label": "Feedback Submissions",
                    "data": [feedback_by_month[m] for m in sorted_months],
                    "borderColor": "#3B82F6",
                    "backgroundColor": "rgba(59, 130, 246, 0.1)",
                    "fill": True
                }],
                "chartType": "line",
                "title": "Feedback Submission Trend"
            },
            "topFeedbackTrainings": top_feedback_trainings
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating feedback analysis report: {str(e)}"
        )

@router.get("/skills-gap-analysis")
async def get_skills_gap_analysis_report(
    db: AsyncSession = Depends(get_db_async),
    current_admin = Depends(get_current_active_admin)
):
    """
    Get comprehensive skills gap analysis
    - Skills with highest gaps
    - Department-wise skill gaps
    - Target vs current expertise distribution
    - Employee skill development progress
    """
    try:
        # Get all competencies
        competencies_result = await db.execute(
            select(EmployeeCompetency)
        )
        competencies = competencies_result.scalars().all()
        
        # Expertise level mapping
        expertise_levels = {
            "L0": 0, "Beginner": 1, "L1": 1,
            "L2": 2, "Intermediate": 2,
            "L3": 3, "Advanced": 3,
            "L4": 4, "Expert": 4,
            "L5": 5
        }
        
        total_skills = len(competencies)
        skills_with_gap = 0
        skills_met = 0
        gaps_by_skill = {}
        gaps_by_dept = {}
        current_expertise_dist = {"Beginner": 0, "Intermediate": 0, "Advanced": 0, "Expert": 0}
        target_expertise_dist = {"Beginner": 0, "Intermediate": 0, "Advanced": 0, "Expert": 0}
        
        for comp in competencies:
            current_level = expertise_levels.get(comp.current_expertise, 0)
            target_level = expertise_levels.get(comp.target_expertise, 0)
            
            # Count gaps
            if current_level < target_level:
                skills_with_gap += 1
                gap = target_level - current_level
                
                # By skill
                skill = comp.skill or "Unknown"
                if skill not in gaps_by_skill:
                    gaps_by_skill[skill] = {"count": 0, "total_gap": 0}
                gaps_by_skill[skill]["count"] += 1
                gaps_by_skill[skill]["total_gap"] += gap
                
                # By department
                dept = comp.department or "Unknown"
                if dept not in gaps_by_dept:
                    gaps_by_dept[dept] = {"count": 0, "total_gap": 0}
                gaps_by_dept[dept]["count"] += 1
                gaps_by_dept[dept]["total_gap"] += gap
            elif current_level >= target_level:
                skills_met += 1
            
            # Distribution
            if comp.current_expertise in ["L0", "L1", "Beginner"]:
                current_expertise_dist["Beginner"] += 1
            elif comp.current_expertise in ["L2", "Intermediate"]:
                current_expertise_dist["Intermediate"] += 1
            elif comp.current_expertise in ["L3", "Advanced"]:
                current_expertise_dist["Advanced"] += 1
            elif comp.current_expertise in ["L4", "L5", "Expert"]:
                current_expertise_dist["Expert"] += 1
            
            if comp.target_expertise in ["L0", "L1", "Beginner"]:
                target_expertise_dist["Beginner"] += 1
            elif comp.target_expertise in ["L2", "Intermediate"]:
                target_expertise_dist["Intermediate"] += 1
            elif comp.target_expertise in ["L3", "Advanced"]:
                target_expertise_dist["Advanced"] += 1
            elif comp.target_expertise in ["L4", "L5", "Expert"]:
                target_expertise_dist["Expert"] += 1
        
        gap_percentage = (skills_with_gap / total_skills * 100) if total_skills > 0 else 0
        
        # Top skills with gaps
        top_gap_skills = []
        for skill, data in gaps_by_skill.items():
            avg_gap = data["total_gap"] / data["count"] if data["count"] > 0 else 0
            top_gap_skills.append({
                "skill": skill,
                "employees_with_gap": data["count"],
                "average_gap": round(avg_gap, 2)
            })
        top_gap_skills.sort(key=lambda x: x["employees_with_gap"], reverse=True)
        top_gap_skills = top_gap_skills[:10]
        
        return {
            "total_count": total_skills,
            "metrics": {
                "total_skills": total_skills,
                "skills_with_gap": skills_with_gap,
                "skills_met": skills_met,
                "gap_percentage": round(gap_percentage, 2),
                "departments_analyzed": len(gaps_by_dept)
            },
            "gapOverviewChart": {
                "labels": ["Skills Met", "Skills with Gap"],
                "datasets": [{
                    "label": "Skills Gap Overview",
                    "data": [skills_met, skills_with_gap],
                    "backgroundColor": ["#10B981", "#EF4444"]
                }],
                "chartType": "pie",
                "title": "Skills Gap Overview"
            },
            "expertiseDistributionChart": {
                "labels": ["Beginner", "Intermediate", "Advanced", "Expert"],
                "datasets": [
                    {
                        "label": "Current Expertise",
                        "data": [
                            current_expertise_dist["Beginner"],
                            current_expertise_dist["Intermediate"],
                            current_expertise_dist["Advanced"],
                            current_expertise_dist["Expert"]
                        ],
                        "backgroundColor": "rgba(59, 130, 246, 0.7)"
                    },
                    {
                        "label": "Target Expertise",
                        "data": [
                            target_expertise_dist["Beginner"],
                            target_expertise_dist["Intermediate"],
                            target_expertise_dist["Advanced"],
                            target_expertise_dist["Expert"]
                        ],
                        "backgroundColor": "rgba(16, 185, 129, 0.7)"
                    }
                ],
                "chartType": "bar",
                "title": "Current vs Target Expertise Distribution"
            },
            "topGapSkills": top_gap_skills
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating skills gap analysis report: {str(e)}"
        )

@router.get("/system-usage-analytics")
async def get_system_usage_analytics(
    db: AsyncSession = Depends(get_db_async),
    current_admin = Depends(get_current_active_admin)
):
    """
    Get system usage analytics
    - Active users trends
    - Training request patterns
    - Assignment submission rates
    - Peak usage times
    """
    try:
        # Get training requests
        requests_result = await db.execute(
            select(TrainingRequest)
        )
        requests = requests_result.scalars().all()
        
        # Get assignment submissions
        submissions_result = await db.execute(
            select(AssignmentSubmission)
        )
        submissions = submissions_result.scalars().all()
        
        # Analyze patterns
        requests_by_status = {"pending": 0, "approved": 0, "rejected": 0}
        requests_by_month = {}
        submissions_by_month = {}
        
        for req in requests:
            status = req.approval_status or "pending"
            requests_by_status[status] = requests_by_status.get(status, 0) + 1
            
            if req.requested_at:
                month_key = req.requested_at.strftime('%Y-%m')
                requests_by_month[month_key] = requests_by_month.get(month_key, 0) + 1
        
        for sub in submissions:
            if sub.submitted_at:
                month_key = sub.submitted_at.strftime('%Y-%m')
                submissions_by_month[month_key] = submissions_by_month.get(month_key, 0) + 1
        
        # Merge months
        all_months = sorted(set(list(requests_by_month.keys()) + list(submissions_by_month.keys())))
        
        return {
            "total_count": len(requests) + len(submissions),
            "metrics": {
                "total_requests": len(requests),
                "total_submissions": len(submissions),
                "pending_requests": requests_by_status["pending"],
                "approved_requests": requests_by_status["approved"],
                "rejected_requests": requests_by_status["rejected"]
            },
            "requestStatusChart": {
                "labels": ["Pending", "Approved", "Rejected"],
                "datasets": [{
                    "label": "Training Requests",
                    "data": [
                        requests_by_status["pending"],
                        requests_by_status["approved"],
                        requests_by_status["rejected"]
                    ],
                    "backgroundColor": ["#F59E0B", "#10B981", "#EF4444"]
                }],
                "chartType": "doughnut",
                "title": "Training Request Status Distribution"
            },
            "activityTrendChart": {
                "labels": all_months,
                "datasets": [
                    {
                        "label": "Training Requests",
                        "data": [requests_by_month.get(m, 0) for m in all_months],
                        "borderColor": "#3B82F6",
                        "backgroundColor": "rgba(59, 130, 246, 0.1)"
                    },
                    {
                        "label": "Assignment Submissions",
                        "data": [submissions_by_month.get(m, 0) for m in all_months],
                        "borderColor": "#10B981",
                        "backgroundColor": "rgba(16, 185, 129, 0.1)"
                    }
                ],
                "chartType": "line",
                "title": "System Activity Trend"
            }
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating system usage analytics: {str(e)}"
        )

@router.get("/export/excel")
async def export_comprehensive_report(
    report_type: str = Query(..., description="Type of report: users, trainings, feedback, skills, all"),
    db: AsyncSession = Depends(get_db_async),
    current_admin = Depends(get_current_active_admin)
):
    """
    Export comprehensive reports as Excel file with multiple sheets
    Supports: users, trainings, feedback, skills, all
    """
    try:
        # Import openpyxl for Excel generation
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"SkillOrbit_Report_{report_type}_{timestamp}.xlsx"
        
        if report_type in ["users", "all"]:
            # Users Sheet
            ws_users = wb.create_sheet("Users")
            users_result = await db.execute(select(User))
            users = users_result.scalars().all()
            
            me_result = await db.execute(select(ManagerEmployee))
            me_relationships = me_result.scalars().all()
            
            # Build user info map
            user_info = {}
            for rel in me_relationships:
                if rel.manager_empid not in user_info:
                    user_info[rel.manager_empid] = {
                        "name": rel.manager_name,
                        "role": "Manager",
                        "is_trainer": rel.manager_is_trainer
                    }
                if rel.employee_empid not in user_info:
                    user_info[rel.employee_empid] = {
                        "name": rel.employee_name,
                        "role": "Employee",
                        "is_trainer": rel.employee_is_trainer
                    }
            
            # Headers
            headers = ["Employee ID", "Name", "Role", "Is Trainer", "Created At"]
            ws_users.append(headers)
            
            # Style headers
            for cell in ws_users[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for user in users:
                info = user_info.get(user.username, {"name": user.username, "role": "Unknown", "is_trainer": False})
                ws_users.append([
                    user.username,
                    info["name"],
                    info["role"],
                    "Yes" if info["is_trainer"] else "No",
                    user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else "N/A"
                ])
            
            # Auto-adjust column widths
            for column in ws_users.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_users.column_dimensions[column_letter].width = adjusted_width
        
        if report_type in ["trainings", "all"]:
            # Trainings Sheet
            ws_trainings = wb.create_sheet("Trainings")
            trainings_result = await db.execute(
                select(TrainingDetail).options(
                    selectinload(TrainingDetail.assignments),
                    selectinload(TrainingDetail.attendance_records)
                )
            )
            trainings = trainings_result.scalars().all()
            
            # Headers
            headers = ["ID", "Training Name", "Trainer", "Skill", "Department", "Division", 
                      "Training Date", "Type", "Assigned", "Attended", "Completion Rate %"]
            ws_trainings.append(headers)
            
            # Style headers
            for cell in ws_trainings[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for training in trainings:
                assigned_count = len(training.assignments)
                attended_count = sum(1 for att in training.attendance_records if att.attended)
                completion_rate = (attended_count / assigned_count * 100) if assigned_count > 0 else 0
                
                ws_trainings.append([
                    training.id,
                    training.training_name,
                    training.trainer_name or "N/A",
                    training.skill or "N/A",
                    training.department or "N/A",
                    training.division or "N/A",
                    training.training_date.strftime('%Y-%m-%d') if training.training_date else "N/A",
                    training.training_type or "N/A",
                    assigned_count,
                    attended_count,
                    round(completion_rate, 2)
                ])
            
            # Auto-adjust column widths
            for column in ws_trainings.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_trainings.column_dimensions[column_letter].width = adjusted_width
        
        if report_type in ["feedback", "all"]:
            # Feedback Sheet
            ws_feedback = wb.create_sheet("Feedback")
            feedback_result = await db.execute(
                select(FeedbackSubmission).options(
                    selectinload(FeedbackSubmission.training)
                )
            )
            feedbacks = feedback_result.scalars().all()
            
            # Headers
            headers = ["Training ID", "Training Name", "Employee ID", "Employee Name", "Submitted At"]
            ws_feedback.append(headers)
            
            # Style headers
            for cell in ws_feedback[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for feedback in feedbacks:
                emp_name = await get_employee_name(db, feedback.employee_empid)
                ws_feedback.append([
                    feedback.training_id,
                    feedback.training.training_name if feedback.training else "N/A",
                    feedback.employee_empid,
                    emp_name,
                    feedback.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if feedback.submitted_at else "N/A"
                ])
            
            # Auto-adjust column widths
            for column in ws_feedback.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_feedback.column_dimensions[column_letter].width = adjusted_width
        
        if report_type in ["skills", "all"]:
            # Skills Gap Sheet
            ws_skills = wb.create_sheet("Skills Gap")
            competencies_result = await db.execute(
                select(EmployeeCompetency)
            )
            competencies = competencies_result.scalars().all()
            
            # Headers
            headers = ["Employee ID", "Employee Name", "Skill", "Competency", "Current Expertise", 
                      "Target Expertise", "Department", "Target Date", "Status"]
            ws_skills.append(headers)
            
            # Style headers
            for cell in ws_skills[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Expertise level mapping
            expertise_levels = {
                "L0": 0, "Beginner": 1, "L1": 1,
                "L2": 2, "Intermediate": 2,
                "L3": 3, "Advanced": 3,
                "L4": 4, "Expert": 4,
                "L5": 5
            }
            
            # Data rows
            for comp in competencies:
                current_level = expertise_levels.get(comp.current_expertise, 0)
                target_level = expertise_levels.get(comp.target_expertise, 0)
                status = "Met" if current_level >= target_level else f"Gap: {target_level - current_level}"
                
                ws_skills.append([
                    comp.employee_empid,
                    comp.employee_name or "N/A",
                    comp.skill or "N/A",
                    comp.competency or "N/A",
                    comp.current_expertise,
                    comp.target_expertise,
                    comp.department or "N/A",
                    comp.target_date.strftime('%Y-%m-%d') if comp.target_date else "N/A",
                    status
                ])
            
            # Auto-adjust column widths
            for column in ws_skills.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_skills.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error exporting Excel report: {str(e)}"
        )
